"""멀티 이슈 파이프라인 dry-run / 등록 루프."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from unittest.mock import MagicMock
from zoneinfo import ZoneInfo

import openpyxl

from ai_work_automation.field_report.excel_ops import FieldIssue, parse_field_issues
from ai_work_automation.field_report.pipeline import load_field_report, run_field_report
from ai_work_automation.opt_in import OptInStore
from ai_work_automation.settings import FieldReportConfig


def _asset_with_issues(tmp_path: Path) -> Path:
    root = tmp_path / "DFS2"
    asset = root / "SDC" / "A6_NX-TSH2326 #1"
    asset.mkdir(parents=True)
    fsr = asset / "[DFS2] 2026 Field Service Report_SDC A6_TSH2326_rev01.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026.08.07"
    ws["T9"] = "00191458"
    ws["V4"] = date(2026, 8, 7)
    ws["V5"] = "이동현"
    ws["L11"] = date(2026, 8, 7)
    from datetime import time

    ws["L12"] = time(9, 30)
    ws["L13"] = time(18, 0)
    ws["B23"] = "□ 이슈 1 (Case : 00191458)"
    ws["B24"] = "□ 이슈 2 (Case : 00196633)"
    ws["B32"] = "작업 종료 후 근무 형태"
    wb.save(fsr)
    wb.close()
    return asset


def test_run_field_report_multi_issue_dry_run(tmp_path: Path) -> None:
    asset = _asset_with_issues(tmp_path)
    plan = load_field_report(asset, sheet_name="2026.08.07")
    tz = ZoneInfo("Asia/Seoul")
    issues = parse_field_issues(
        plan.workbook,
        plan.sheet_name,
        day=date(2026, 8, 7),
        fse_name="이동현",
        work_start=datetime(2026, 8, 7, 9, 30, tzinfo=tz),
        work_end=datetime(2026, 8, 7, 18, 0, tzinfo=tz),
    )
    issues[0].case_id = "500AAA"
    issues[1].case_id = "500BBB"
    sf = MagicMock()
    sf.find_technical_service_wos_on_day.return_value = []
    opt = OptInStore(tmp_path / "opt.json")
    result = run_field_report(
        plan,
        sf=sf,
        opt_in=opt,
        case_ids=[],
        dry_run=True,
        cfg=FieldReportConfig(),
        issues=issues,
    )
    assert result.status == "dry_run"
    would = (result.details or {})["would"]
    assert len(would["issues"]) == 2
    assert would["issues"][0]["activity_line"] != would["issues"][1]["activity_line"]
    assert would["issues"][0]["start_date"] != would["issues"][1]["start_date"]
    sf.append_case_activities.assert_not_called()


def test_run_field_report_multi_issue_real(tmp_path: Path) -> None:
    asset = _asset_with_issues(tmp_path)
    plan = load_field_report(asset, sheet_name="2026.08.07")
    issues = [
        FieldIssue(
            case_number="00191458",
            issue_line="□ 이슈 1 (Case : 00191458)",
            activity_line="2026-08-07 [이동현] □ 이슈 1 (Case : 00191458)",
            start=datetime(2026, 8, 7, 9, 30, tzinfo=ZoneInfo("Asia/Seoul")),
            end=datetime(2026, 8, 7, 13, 45, tzinfo=ZoneInfo("Asia/Seoul")),
            included=True,
            case_id="500AAA",
        ),
        FieldIssue(
            case_number="00196633",
            issue_line="□ 이슈 2 (Case : 00196633)",
            activity_line="2026-08-07 [이동현] □ 이슈 2 (Case : 00196633)",
            start=datetime(2026, 8, 7, 13, 45, tzinfo=ZoneInfo("Asia/Seoul")),
            end=datetime(2026, 8, 7, 18, 0, tzinfo=ZoneInfo("Asia/Seoul")),
            included=True,
            case_id="500BBB",
        ),
    ]
    sf = MagicMock()
    sf.find_technical_service_wos_on_day.return_value = []
    sf.create_technical_service_work_order.side_effect = ["0WO1", "0WO2"]
    sf.attach_file_to_record.side_effect = ["068a", "068b"]
    opt = OptInStore(tmp_path / "opt.json")
    result = run_field_report(
        plan,
        sf=sf,
        opt_in=opt,
        case_ids=[],
        dry_run=False,
        cfg=FieldReportConfig(),
        issues=issues,
    )
    assert result.status == "success"
    assert sf.append_case_activities.call_count == 2
    assert sf.append_case_activities.call_args_list[0].kwargs.get("enforce_cutoff") is False
    assert sf.append_case_activities.call_args_list[0].args[1].endswith("이슈 1 (Case : 00191458)")
    assert sf.append_case_activities.call_args_list[1].args[1].endswith("이슈 2 (Case : 00196633)")
    assert sf.create_technical_service_work_order.call_count == 2
    desc0 = sf.create_technical_service_work_order.call_args_list[0].kwargs["description"]
    act0 = sf.append_case_activities.call_args_list[0].args[1]
    assert desc0 != act0
    assert "이슈 1" in desc0
    starts = [
        c.kwargs["start_date"] for c in sf.create_technical_service_work_order.call_args_list
    ]
    assert starts[0] != starts[1]


def test_run_field_report_continues_after_one_failure(tmp_path: Path) -> None:
    asset = _asset_with_issues(tmp_path)
    plan = load_field_report(asset, sheet_name="2026.08.07")
    issues = [
        FieldIssue(
            case_number="00191458",
            issue_line="□ 이슈 1 (Case : 00191458)",
            activity_line="2026-08-07 [이동현] □ 이슈 1 (Case : 00191458)",
            detail_text="상세 A",
            start=datetime(2026, 8, 7, 9, 30, tzinfo=ZoneInfo("Asia/Seoul")),
            end=datetime(2026, 8, 7, 13, 45, tzinfo=ZoneInfo("Asia/Seoul")),
            included=True,
            case_id="500AAA",
        ),
        FieldIssue(
            case_number="00150143",
            issue_line="□ 이슈 2 (Case : 00150143)",
            activity_line="2026-08-07 [이동현] □ 이슈 2 (Case : 00150143)",
            detail_text="상세 B",
            start=datetime(2026, 8, 7, 13, 45, tzinfo=ZoneInfo("Asia/Seoul")),
            end=datetime(2026, 8, 7, 18, 0, tzinfo=ZoneInfo("Asia/Seoul")),
            included=True,
            case_id="500BBB",
        ),
    ]
    sf = MagicMock()
    sf.find_technical_service_wos_on_day.return_value = []

    def _append(cid, line, **kwargs):
        if cid == "500AAA":
            raise RuntimeError("boom")

    sf.append_case_activities.side_effect = _append
    sf.create_technical_service_work_order.return_value = "0WO2"
    sf.attach_file_to_record.return_value = "068b"
    opt = OptInStore(tmp_path / "opt.json")
    result = run_field_report(
        plan,
        sf=sf,
        opt_in=opt,
        case_ids=[],
        dry_run=False,
        cfg=FieldReportConfig(),
        issues=issues,
    )
    assert result.status == "partial"
    assert len((result.details or {})["acted"]) == 1
    assert (result.details or {})["acted"][0]["case_id"] == "500BBB"
    assert len((result.details or {})["failed"]) == 1
