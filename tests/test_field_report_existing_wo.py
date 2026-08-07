"""이미 등록된 Technical Service WO 감지·스킵."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from unittest.mock import MagicMock
from zoneinfo import ZoneInfo

import openpyxl

from ai_work_automation.field_report.excel_ops import FieldIssue
from ai_work_automation.field_report.pipeline import load_field_report, run_field_report
from ai_work_automation.opt_in import OptInStore
from ai_work_automation.settings import FieldReportConfig
from ai_work_automation.sf.adapter import SalesforceAdapter, start_date_matches_day


def test_start_date_matches_day() -> None:
    assert start_date_matches_day("2026-08-07T09:30:00.000+0900", date(2026, 8, 7))
    assert start_date_matches_day("2026-08-07T13:45:00.000+0900", date(2026, 8, 7))
    assert not start_date_matches_day("2026-08-06T18:00:00.000+0900", date(2026, 8, 7))
    assert not start_date_matches_day(None, date(2026, 8, 7))


def test_find_technical_service_wos_on_day_filters() -> None:
    client = MagicMock()
    client.query.return_value = {
        "records": [
            {
                "Id": "0WOold",
                "WorkOrderNumber": "00026050",
                "StartDate": "2026-08-06T09:00:00.000+0900",
                "RecordTypeId": "0120o000001lQJ5AAM",
            },
            {
                "Id": "0WOtoday",
                "WorkOrderNumber": "00026058",
                "StartDate": "2026-08-07T09:30:00.000+0900",
                "RecordTypeId": "0120o000001lQJ5AAM",
            },
        ]
    }
    adapter = SalesforceAdapter(
        client,
        cutoff=datetime(2025, 1, 1, tzinfo=ZoneInfo("UTC")),
        technical_service_record_type_id="0120o000001lQJ5AAM",
    )
    found = adapter.find_technical_service_wos_on_day("500xx", date(2026, 8, 7))
    assert len(found) == 1
    assert found[0].id == "0WOtoday"
    assert found[0].work_order_number == "00026058"


def _asset(tmp_path: Path) -> Path:
    asset = tmp_path / "SDC" / "A6_NX-TSH2326 #1"
    asset.mkdir(parents=True)
    fsr = asset / "[DFS2] 2026 Field Service Report_SDC A6.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026.08.07"
    ws["T9"] = "00191458"
    ws["V4"] = date(2026, 8, 7)
    ws["V5"] = "이동현"
    ws["B23"] = "□ 이슈 1 (Case : 00191458)"
    ws["B32"] = "작업 종료 후 근무 형태"
    wb.save(fsr)
    wb.close()
    return asset


def test_run_skips_existing_wo_no_duplicate_create(tmp_path: Path) -> None:
    asset = _asset(tmp_path)
    plan = load_field_report(asset, sheet_name="2026.08.07")
    issues = [
        FieldIssue(
            case_number="00191458",
            issue_line="□ 이슈 1 (Case : 00191458)",
            activity_line="2026-08-07 [이동현] □ 이슈 1",
            start=datetime(2026, 8, 7, 9, 30, tzinfo=ZoneInfo("Asia/Seoul")),
            end=datetime(2026, 8, 7, 18, 0, tzinfo=ZoneInfo("Asia/Seoul")),
            included=True,
            case_id="500AAA",
        )
    ]
    existing = MagicMock()
    existing.id = "0WOexist"
    existing.work_order_number = "00026058"
    existing.case_id = "500AAA"

    sf = MagicMock()
    sf.find_technical_service_wos_on_day.return_value = [existing]
    opt = OptInStore(tmp_path / "opt.json")
    result = run_field_report(
        plan,
        sf=sf,
        opt_in=opt,
        case_ids=[],
        dry_run=False,
        cfg=FieldReportConfig(),
        issues=issues,
    )
    assert result.status == "success"
    sf.create_technical_service_work_order.assert_not_called()
    sf.append_case_activities.assert_not_called()
    sf.attach_file_to_record.assert_not_called()
    acted = (result.details or {})["acted"]
    assert acted[0]["skipped"] is True
    assert acted[0]["work_order_number"] == "00026058"
    assert (result.details or {}).get("skipped_existing")
