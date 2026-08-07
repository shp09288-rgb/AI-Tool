from datetime import date, time
from pathlib import Path

import openpyxl

from ai_work_automation.field_report.excel_ops import (
    DaySheetMeta,
    _as_time,
    ensure_day_sheet,
    export_sheet_workbook,
    find_report_crop_end_row,
    find_report_workbook,
    find_time_overlaps,
    list_asset_folders,
    list_day_sheets,
    parse_field_issues,
    read_day_sheet_meta,
    report_crop_a1_range,
    resolve_report_mode,
    sheet_to_html,
    split_workday_slots,
)
from openpyxl.styles import Font, PatternFill
from ai_work_automation.field_report.pipeline import load_field_report, prepare_field_report
import pytest


def _make_fsr(path: Path) -> None:
    from datetime import time

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026.07.07"
    ws["R9"] = "CRM Case ID"
    ws["T9"] = "00196633"
    ws["V4"] = date(2026, 7, 7)
    ws["V5"] = "이동현"
    ws["L11"] = date(2026, 7, 7)
    ws["L12"] = time(9, 30)
    ws["L13"] = time(15, 0)
    ws["B23"] = "Tip 교체"
    ws["B28"] = "Next Plan Update"
    ws["B30"] = "Data Location"
    # 첨부/미리보기에서 제외해야 하는 하단 근무형태 블록
    ws["B32"] = "작업 종료 후 근무 형태"
    ws["C32"] = "업무 종료"
    ws["B33"] = "재택 근무 전환 시"
    ws["B28"].fill = PatternFill("solid", fgColor="538DD5")
    ws["B28"].font = Font(bold=True, color="FFFFFF")
    wb.create_sheet("SW Version history")
    wb.save(path)
    wb.close()


def test_as_time_parses_ampm_strings() -> None:
    assert _as_time("18:00 PM") == time(18, 0)
    assert _as_time("9:30 AM") == time(9, 30)
    assert _as_time("12:00 PM") == time(12, 0)
    assert _as_time("12:00 AM") == time(0, 0)
    assert _as_time("18:00") == time(18, 0)
    assert _as_time(time(15, 0)) == time(15, 0)


def test_end_datetime_none_when_end_time_missing() -> None:
    meta = DaySheetMeta(
        case_numbers=[],
        fse_name="x",
        report_date=date(2026, 8, 7),
        end_time=None,
    )
    assert meta.end_datetime() is None


def test_read_meta_parses_end_time_ampm(tmp_path: Path) -> None:
    fsr = tmp_path / "report.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026.08.07"
    ws["T9"] = "00191458"
    ws["V4"] = date(2026, 8, 7)
    ws["V5"] = "이동현"
    ws["L11"] = date(2026, 8, 7)
    ws["L12"] = time(9, 30)
    ws["L13"] = "18:00 PM"
    wb.save(fsr)
    wb.close()
    meta = read_day_sheet_meta(fsr, "2026.08.07")
    assert meta.end_time == time(18, 0)
    assert meta.end_datetime() is not None
    assert meta.end_datetime().hour == 18


def test_crop_excludes_post_work_block(tmp_path: Path) -> None:
    fsr = tmp_path / "report.xlsx"
    _make_fsr(fsr)
    wb = openpyxl.load_workbook(fsr)
    ws = wb["2026.07.07"]
    end = find_report_crop_end_row(ws)
    a1, end_row, end_col = report_crop_a1_range(ws)
    wb.close()
    assert end == 31  # B32 마커 직전 행까지
    assert end_row == 31
    assert a1.startswith("A1:") and a1.endswith("31")

    html = sheet_to_html(fsr, "2026.07.07")
    assert "Tip 교체" in html
    assert "Next Plan Update" in html
    assert "작업 종료 후 근무 형태" not in html
    assert "재택 근무 전환 시" not in html
    assert "업무 종료" not in html
    assert "#538dd5" in html.lower()  # 헤더 배경색 반영

    out = tmp_path / "day.xlsx"
    export_sheet_workbook(fsr, "2026.07.07", out)
    exported = openpyxl.load_workbook(out)
    ws = exported.active
    assert ws["B23"].value == "Tip 교체"
    assert ws["B28"].value == "Next Plan Update"
    assert ws["B32"].value is None
    assert "작업 종료 후 근무 형태" not in [
        str(c.value) for row in ws.iter_rows(max_row=40, max_col=5) for c in row if c.value
    ]
    exported.close()


def test_list_asset_folders_and_find_fsr(tmp_path: Path) -> None:
    root = tmp_path / "DFS2"
    asset = root / "SDC" / "A6_NX-TSH2326 #1"
    asset.mkdir(parents=True)
    fsr = asset / "[DFS2] 2026 Field Service Report_SDC A6_TSH2326_rev01.xlsx"
    _make_fsr(fsr)
    folders = list_asset_folders(root, customer="SDC")
    assert any(p.name == "A6_NX-TSH2326 #1" for p in folders)
    mode = resolve_report_mode(asset)
    assert mode == "field_service"
    found = find_report_workbook(asset, mode="field_service")
    assert found == fsr


def test_ensure_day_sheet_and_meta_and_export(tmp_path: Path) -> None:
    fsr = tmp_path / "report.xlsx"
    _make_fsr(fsr)
    sheet_name, created = ensure_day_sheet(fsr, day=date(2026, 8, 7))
    assert sheet_name == "2026.08.07"
    assert created is True
    meta = read_day_sheet_meta(fsr, sheet_name)
    # 새 시트는 이전 출장 본문·Case ID를 비움
    assert meta.case_numbers == []
    assert meta.fse_name == "이동현"
    assert meta.start_time.hour == 9 and meta.start_time.minute == 30
    assert meta.end_time.hour == 15
    assert meta.start_datetime() is not None
    wb_check = openpyxl.load_workbook(fsr)
    assert wb_check["2026.08.07"]["B23"].value is None
    assert wb_check["2026.08.07"]["T9"].value is None
    wb_check.close()
    assert "2026.08.07" in sheet_to_html(fsr, sheet_name)
    out = tmp_path / "day.xlsx"
    export_sheet_workbook(fsr, sheet_name, out)
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["2026.08.07"]
    assert wb.active["T9"].value is None
    wb.close()


def test_list_day_sheets_accepts_hyphen_dates(tmp_path: Path) -> None:
    fsr = tmp_path / "report.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "2026.07.09"
    wb.create_sheet("2026-08-07")
    wb.save(fsr)
    wb.close()
    sheets = list_day_sheets(fsr)
    assert "2026-08-07" in sheets
    assert sheets[0] == "2026-08-07"  # 최신 날짜 우선


def test_list_day_sheets_mmdd_with_year_file(tmp_path: Path) -> None:
    from ai_work_automation.field_report.excel_ops import (
        list_report_workbooks,
        year_hint_from_path,
    )

    asset = tmp_path / "SDC" / "A5_NX-TSH2225 #1"
    asset.mkdir(parents=True)
    fsr = asset / "2026_[Field Service Report]_SDC A5_NX-TSH2225 #1.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "SW Version history"
    wb.create_sheet("0729")
    wb.create_sheet("0807")
    wb.create_sheet("1203(유상)")
    wb.create_sheet("출장준비 Check")
    wb.save(fsr)
    wb.close()
    assert year_hint_from_path(fsr) == 2026
    sheets = list_day_sheets(fsr)
    assert sheets[0] == "1203(유상)"
    assert "0807" in sheets and "0729" in sheets
    wbs = list_report_workbooks(asset, mode="field_service")
    assert wbs[0] == fsr


def test_list_day_sheets_reads_via_temp_copy_when_locked(tmp_path: Path, monkeypatch) -> None:
    """Excel 잠금(PermissionError) 시 임시 복사본으로 읽는지."""
    import ai_work_automation.field_report.excel_ops as ops

    fsr = tmp_path / "2026_report.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "0729"
    wb.save(fsr)
    wb.close()

    real_load = openpyxl.load_workbook
    calls = {"n": 0}

    def flaky_load(path, **kwargs):
        calls["n"] += 1
        if calls["n"] == 1:
            raise PermissionError("locked")
        return real_load(path, **kwargs)

    monkeypatch.setattr(ops.openpyxl, "load_workbook", flaky_load)
    sheets = list_day_sheets(fsr)
    assert "0729" in sheets
    assert calls["n"] >= 2


def test_load_field_report_reads_existing_sheet_only(tmp_path: Path) -> None:
    root = tmp_path / "DFS2"
    asset = root / "SDC" / "A6_NX-TSH2326 #1"
    asset.mkdir(parents=True)
    fsr = asset / "[DFS2] 2026 Field Service Report_SDC A6_TSH2326_rev01.xlsx"
    _make_fsr(fsr)
    assert list_day_sheets(fsr) == ["2026.07.07"]
    plan = load_field_report(asset, sheet_name="2026.07.07")
    assert plan.sheet_name == "2026.07.07"
    assert plan.sheet_created is False
    assert plan.meta.case_numbers == ["00196633"]
    with pytest.raises(FileNotFoundError, match="시트가 없습니다"):
        prepare_field_report(asset, day=date(2026, 8, 7))


def test_ensure_day_sheet_recreate_clears_stale_content(tmp_path: Path) -> None:
    fsr = tmp_path / "report.xlsx"
    _make_fsr(fsr)
    ensure_day_sheet(fsr, day=date(2026, 8, 7))
    # 잘못 남은 본문 시뮬레이션
    wb = openpyxl.load_workbook(fsr)
    wb["2026.08.07"]["B23"] = "옛 출장 내용"
    wb["2026.08.07"]["T9"] = "00197302"
    wb.save(fsr)
    wb.close()
    _, recreated = ensure_day_sheet(fsr, day=date(2026, 8, 7), recreate=True)
    assert recreated is True
    meta = read_day_sheet_meta(fsr, "2026.08.07")
    assert meta.case_numbers == []
    wb2 = openpyxl.load_workbook(fsr)
    assert wb2["2026.08.07"]["B23"].value is None
    wb2.close()


def test_split_workday_slots_equal() -> None:
    from datetime import datetime
    from zoneinfo import ZoneInfo

    tz = ZoneInfo("Asia/Seoul")
    start = datetime(2026, 8, 7, 9, 30, tzinfo=tz)
    end = datetime(2026, 8, 7, 18, 0, tzinfo=tz)
    assert split_workday_slots(start, end, 1) == [(start, end)]
    slots = split_workday_slots(start, end, 3)
    assert len(slots) == 3
    assert slots[0][0] == start
    assert slots[-1][1] == end
    assert slots[0][1] == slots[1][0]
    assert slots[1][1] == slots[2][0]
    assert find_time_overlaps([]) == []


def test_parse_field_issues_multi_and_fallback(tmp_path: Path) -> None:
    from datetime import datetime
    from zoneinfo import ZoneInfo

    fsr = tmp_path / "report.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026.08.07"
    ws["B23"] = "□ 이슈 1 (Case : 00191458)"
    ws["B24"] = "□ 이슈 2 (Case：00196633)"  # 전각 콜론
    ws["B25"] = "이슈 3 (Case : 00197302)"  # 체크 없음 보조 매칭
    ws["B28"] = "Next Plan Update"
    ws["B32"] = "작업 종료 후 근무 형태"
    wb.save(fsr)
    wb.close()

    tz = ZoneInfo("Asia/Seoul")
    work_start = datetime(2026, 8, 7, 9, 30, tzinfo=tz)
    work_end = datetime(2026, 8, 7, 18, 0, tzinfo=tz)
    issues = parse_field_issues(
        fsr,
        "2026.08.07",
        day=date(2026, 8, 7),
        fse_name="이동현",
        work_start=work_start,
        work_end=work_end,
    )
    assert len(issues) == 3
    assert [i.case_number for i in issues] == ["00191458", "00196633", "00197302"]
    assert issues[0].activity_line.startswith("2026-08-07 [이동현] □ 이슈 1")
    assert issues[0].start == work_start
    assert issues[-1].end == work_end
    assert issues[0].end == issues[1].start
    assert find_time_overlaps(issues) == []

    # 패턴 없음
    wb2 = openpyxl.Workbook()
    wb2.active.title = "2026.08.07"
    wb2.active["B23"] = "일반 메모만"
    wb2.save(fsr)
    wb2.close()
    assert (
        parse_field_issues(
            fsr,
            "2026.08.07",
            day=date(2026, 8, 7),
            fse_name="이동현",
            work_start=work_start,
            work_end=work_end,
        )
        == []
    )
