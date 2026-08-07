"""로컬 DFS2 Field Service / Installation 엑셀 조작."""

from __future__ import annotations

import re
import shutil
import tempfile
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from html import escape
from pathlib import Path
from typing import Iterator
from xml.etree import ElementTree as ET
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# 실무: "2026.07.09" / "2026-08-07" / 연도파일+시트 "0729" / "1203(유상)"
_DATE_SHEET_RE = re.compile(r"^(\d{4})[.\-](\d{2})[.\-](\d{2})")
# MMDD 단독 또는 접미사 (1203(유상), 1120_유상 건, 1105~6)
_MMDD_SHEET_RE = re.compile(r"^(\d{2})(\d{2})(?:$|[^0-9])")
_YEAR_IN_NAME_RE = re.compile(r"(?:^|[_\s\[])(\d{4})(?:[_\s\]]|$)")
_INSTALL_DIR = "Installation 레포트_백업용"
_FSR_GLOB = "*Field Service Report*.xlsx"
_INSTALL_GLOB = "*Installation Report*.xlsx"
_SEOUL = ZoneInfo("Asia/Seoul")
# 첨부·메일·미리보기에서 제외할 하단 블록 시작 표식
_CROP_END_MARKER = "작업 종료 후 근무 형태"
_DEFAULT_CROP_END_ROW = 31
_PREVIEW_MAX_COL = 24  # A~X (보고서 본문 폭)
_DRAWINGML_NS = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
# Office 기본 테마 (파일에 theme가 없을 때)
_DEFAULT_THEME_RGB = (
    "000000",  # 0 dk1
    "FFFFFF",  # 1 lt1
    "1F497D",  # 2 dk2
    "EEECE1",  # 3 lt2
    "4F81BD",  # 4 accent1
    "C0504D",  # 5 accent2
    "9BBB59",  # 6 accent3
    "8064A2",  # 7 accent4
    "4BACC6",  # 8 accent5
    "F79646",  # 9 accent6
    "0000FF",  # 10 hlink
    "800080",  # 11 folHlink
)


@dataclass
class DaySheetMeta:
    case_numbers: list[str]
    fse_name: str
    report_date: date | None
    summary_hint: str = ""
    start_date: date | None = None
    start_time: time | None = None
    end_time: time | None = None

    def start_datetime(self) -> datetime | None:
        return _combine_local(self.start_date or self.report_date, self.start_time)

    def end_datetime(self) -> datetime | None:
        # end_time 없으면 midnight으로 꾸미지 않음 (UI가 00:00으로 보이는 원인)
        if self.end_time is None:
            return None
        return _combine_local(self.start_date or self.report_date, self.end_time)


@dataclass
class FieldIssue:
    """Daily Note에서 파싱한 이슈(=Case) 한 줄."""

    case_number: str
    issue_line: str
    activity_line: str
    detail_text: str = ""  # WO Description용 상세 (이슈 줄 아래 본문)
    start: datetime | None = None
    end: datetime | None = None
    included: bool = True
    case_id: str | None = None  # SF Id (UI에서 resolve)


_ISSUE_LINE_RE = re.compile(
    r"^[□☐]?\s*(?P<label>.+?)\s*\(\s*Case\s*[:：]\s*(?P<case>\d{8})\s*\)\s*$",
    re.IGNORECASE,
)


def _as_date(raw: object) -> date | None:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    return None


def _as_time(raw: object) -> time | None:
    if isinstance(raw, datetime):
        return raw.time().replace(microsecond=0)
    if isinstance(raw, time):
        return raw.replace(microsecond=0)
    if isinstance(raw, str) and raw.strip():
        text = re.sub(r"\s+", " ", raw.strip())
        # 실무 시트: "18:00 PM", "9:30 AM" (24시+PM 혼용도 허용)
        m = re.fullmatch(
            r"(?P<h>\d{1,2}):(?P<min>\d{2})(?::(?P<sec>\d{2}))?\s*(?P<ampm>[AaPp][Mm])?",
            text,
        )
        if m:
            hour = int(m.group("h"))
            minute = int(m.group("min"))
            second = int(m.group("sec") or 0)
            ampm = (m.group("ampm") or "").upper()
            if ampm:
                if hour < 1 or hour > 12:
                    # "18:00 PM" 처럼 이미 24시인 경우 AM/PM 무시
                    if hour > 23:
                        return None
                elif ampm == "AM":
                    hour = 0 if hour == 12 else hour
                else:  # PM
                    hour = hour if hour == 12 else hour + 12
            if 0 <= hour <= 23 and 0 <= minute <= 59 and 0 <= second <= 59:
                return time(hour, minute, second)
        for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p"):
            try:
                return datetime.strptime(text, fmt).time()
            except ValueError:
                continue
    return None


def _combine_local(d: date | None, t: time | None) -> datetime | None:
    if d is None or t is None:
        return None
    return datetime.combine(d, t, tzinfo=_SEOUL)


def to_sf_datetime(dt: datetime | None) -> str | None:
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=_SEOUL)
    return dt.strftime("%Y-%m-%dT%H:%M:%S.000%z")


def list_customers(root: Path) -> list[str]:
    if not root.is_dir():
        return []
    return sorted(
        p.name
        for p in root.iterdir()
        if p.is_dir() and not p.name.startswith("(") and p.name != "DFS2 Customer Management"
    )


def list_asset_folders(root: Path, customer: str) -> list[Path]:
    base = root / customer
    if not base.is_dir():
        return []
    return sorted([p for p in base.iterdir() if p.is_dir()], key=lambda p: p.name)


def resolve_report_mode(asset_dir: Path) -> str:
    """field_service | installation — FSR 파일이 있으면 field_service."""
    if find_report_workbook(asset_dir, mode="field_service") is not None:
        return "field_service"
    return "installation"


@contextmanager
def load_workbook_resilient(
    workbook_path: Path, **kwargs
) -> Iterator[Workbook]:
    """Excel이 파일을 연 상태(PermissionError)여도 임시 복사본으로 읽는다."""
    path = Path(workbook_path)
    tmp_dir: Path | None = None
    try:
        try:
            wb = openpyxl.load_workbook(path, **kwargs)
        except PermissionError:
            tmp_dir = Path(tempfile.mkdtemp(prefix="fr_xlsx_"))
            tmp_path = tmp_dir / path.name
            shutil.copy2(path, tmp_path)
            wb = openpyxl.load_workbook(tmp_path, **kwargs)
        try:
            yield wb
        finally:
            wb.close()
    finally:
        if tmp_dir is not None:
            shutil.rmtree(tmp_dir, ignore_errors=True)


def year_hint_from_path(path: Path) -> int | None:
    """파일명에서 연도 추정 (예: 2026_[Field Service Report]_....xlsx)."""
    m = re.match(r"^(\d{4})[_\[ ]", path.stem)
    if m:
        return int(m.group(1))
    found = _YEAR_IN_NAME_RE.findall(path.stem)
    if found:
        return int(found[0])
    return None


def list_report_workbooks(asset_dir: Path, *, mode: str) -> list[Path]:
    """설비 폴더의 리포트 xlsx 목록 (최신 수정·현재 연도 파일 우선)."""
    if mode == "field_service":
        found = list(asset_dir.glob(_FSR_GLOB))
    else:
        found = []
        for r in (asset_dir / _INSTALL_DIR, asset_dir):
            if r.is_dir():
                found.extend(r.glob(_INSTALL_GLOB))
    this_year = date.today().year

    def _rank(p: Path) -> tuple:
        y = year_hint_from_path(p)
        # 현재 연도 파일을 앞에, 그다음 mtime
        year_rank = 0 if y == this_year else (1 if y else 2)
        return (year_rank, -p.stat().st_mtime, p.name.lower())

    return sorted(found, key=_rank)


def find_report_workbook(asset_dir: Path, *, mode: str) -> Path | None:
    matches = list_report_workbooks(asset_dir, mode=mode)
    return matches[0] if matches else None


def _sheet_sort_key(name: str, year_hint: int) -> str:
    m = _DATE_SHEET_RE.match(name)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = _MMDD_SHEET_RE.match(name)
    if m:
        return f"{year_hint:04d}-{m.group(1)}-{m.group(2)}"
    return ""


def _dated_sheet_names(wb: openpyxl.Workbook, *, year_hint: int | None = None) -> list[str]:
    year = year_hint or date.today().year
    names: list[str] = []
    for n in wb.sheetnames:
        if _DATE_SHEET_RE.match(n):
            names.append(n)
            continue
        m = _MMDD_SHEET_RE.match(n)
        if not m:
            continue
        try:
            date(year, int(m.group(1)), int(m.group(2)))
        except ValueError:
            continue
        names.append(n)
    return sorted(names, key=lambda n: _sheet_sort_key(n, year), reverse=True)


def sheet_name_for_day(day: date) -> str:
    """기본 표기(점). 매칭 시에는 sheet_name_aliases_for_day 사용."""
    return day.strftime("%Y.%m.%d")


def sheet_name_aliases_for_day(day: date) -> list[str]:
    """같은 날짜의 가능한 시트명 (점 / 하이픈 / MMDD)."""
    return [
        day.strftime("%Y.%m.%d"),
        day.strftime("%Y-%m-%d"),
        day.strftime("%m%d"),
    ]


def list_day_sheets(
    workbook_path: Path, *, year_hint: int | None = None
) -> list[str]:
    """워크북의 일자 시트 이름 (최신 날짜 우선). MMDD 시트는 year_hint/파일명 연도 사용."""
    year = year_hint if year_hint is not None else year_hint_from_path(workbook_path)
    with load_workbook_resilient(
        workbook_path, read_only=True, data_only=True
    ) as wb:
        return _dated_sheet_names(wb, year_hint=year)


def _is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _blank_new_day_content(ws: Worksheet, day: date) -> None:
    """서식은 유지하고, 이전 출장 본문·Case ID 등 일자 고유 값만 비운다."""
    ws["V4"] = day
    if not _is_formula(ws["L11"].value):
        ws["L11"] = day
    # Case / 작업 본문 (이전 일자 내용이 첨부되지 않도록)
    ws["T9"] = None
    # 본문 영역: FSR 양식상 작업 내용은 대략 23행 이하
    for row in range(23, 81):
        for col in range(2, 16):  # B–O
            cell = ws.cell(row, col)
            if cell.value is not None and not _is_formula(cell.value):
                cell.value = None


def ensure_day_sheet(
    workbook_path: Path,
    *,
    day: date | None = None,
    recreate: bool = False,
) -> tuple[str, bool]:
    """지정일 시트를 준비한다.

    - 없으면: 가장 최근 일자 시트 서식을 복사한 뒤 본문을 비우고 날짜를 넣는다.
    - 있고 recreate=True: 기존 시트를 지우고 위와 같이 다시 만든다.
    - 있고 recreate=False: 기존 시트를 그대로 쓴다.

    Returns:
        (sheet_name, created_or_recreated)
    """
    day = day or date.today()
    target = sheet_name_for_day(day)
    try:
        wb = openpyxl.load_workbook(workbook_path)
    except PermissionError as exc:
        raise PermissionError(
            f"엑셀이 열려 있거나 OneDrive가 잠근 상태입니다. 파일을 닫고 다시 시도하세요: {workbook_path}"
        ) from exc
    try:
        if target in wb.sheetnames:
            if not recreate:
                return target, False
            del wb[target]

        dated = [
            n
            for n in _dated_sheet_names(wb, year_hint=year_hint_from_path(workbook_path))
            if n != target
        ]
        if not dated:
            raise ValueError("복사할 일자 시트가 없습니다")
        # 자기 자신(방금 지운 시트)이 아닌 최신 서식 소스
        source = wb[dated[0]]
        new_ws = wb.copy_worksheet(source)
        new_ws.title = target
        _blank_new_day_content(new_ws, day)
        # 탭 목록 맨 앞으로 옮겨 Excel에서 바로 보이게
        try:
            wb.move_sheet(new_ws, offset=-wb.sheetnames.index(target))
        except Exception:  # noqa: BLE001 — 위치 이동 실패해도 생성은 유지
            pass
        try:
            wb.save(workbook_path)
        except PermissionError as exc:
            raise PermissionError(
                f"엑셀 저장 실패(파일 잠금). Excel/Teams에서 닫은 뒤 다시 시도하세요: {workbook_path}"
            ) from exc
        return target, True
    finally:
        wb.close()


def _parse_case_numbers(raw: object) -> list[str]:
    if raw is None:
        return []
    text = str(raw).replace("\x00", "").strip()
    if not text:
        return []
    parts = re.split(r"[,/\s]+", text)
    out: list[str] = []
    for p in parts:
        p = p.strip()
        if re.fullmatch(r"\d{8}", p):
            out.append(p)
    return out


def read_day_sheet_meta(
    workbook_path: Path,
    sheet_name: str,
    *,
    case_id_cell: str = "T9",
    fse_name_cell: str = "V5",
    report_date_cell: str = "V4",
    summary_cell: str = "B23",
    start_date_cell: str = "L11",
    start_time_cell: str = "L12",
    end_time_cell: str = "L13",
) -> DaySheetMeta:
    with load_workbook_resilient(workbook_path, data_only=True) as wb:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"시트 없음: {sheet_name}")
        ws: Worksheet = wb[sheet_name]
        # 템플릿마다 Case ID 위치가 다름 (T9 / X9 등)
        case_numbers = _parse_case_numbers(ws[case_id_cell].value)
        if not case_numbers:
            for alt in ("X9", "T9", "W9", "U9"):
                if alt == case_id_cell:
                    continue
                case_numbers = _parse_case_numbers(ws[alt].value)
                if case_numbers:
                    break
        fse = str(ws[fse_name_cell].value or "").replace("\x00", "").strip()
        report_date = _as_date(ws[report_date_cell].value)
        start_date = _as_date(ws[start_date_cell].value) or report_date
        summary = str(ws[summary_cell].value or "").replace("\x00", "").strip()
        if len(summary) > 120:
            summary = summary[:117] + "..."
        return DaySheetMeta(
            case_numbers=case_numbers,
            fse_name=fse,
            report_date=report_date,
            summary_hint=summary,
            start_date=start_date,
            start_time=_as_time(ws[start_time_cell].value),
            end_time=_as_time(ws[end_time_cell].value),
        )


def find_report_crop_end_row(ws: Worksheet) -> int:
    """첨부·미리보기에 포함할 마지막 행 번호.

    '작업 종료 후 근무 형태' 행 직전까지. 마커가 없으면 기본값.
    """
    max_scan = min(ws.max_row or _DEFAULT_CROP_END_ROW, 80)
    max_col = min(ws.max_column or 5, 10)
    for row in ws.iter_rows(min_row=1, max_row=max_scan, max_col=max_col):
        for cell in row:
            val = cell.value
            if val is not None and _CROP_END_MARKER in str(val):
                return max(1, cell.row - 1)
    return _DEFAULT_CROP_END_ROW


def _parse_theme_rgbs(wb) -> tuple[str, ...]:
    raw = getattr(wb, "loaded_theme", None)
    if not raw:
        return _DEFAULT_THEME_RGB
    try:
        root = ET.fromstring(raw)
    except ET.ParseError:
        return _DEFAULT_THEME_RGB
    scheme = root.find(".//a:clrScheme", _DRAWINGML_NS)
    if scheme is None:
        return _DEFAULT_THEME_RGB
    colors: list[str] = []
    for child in list(scheme):
        srgb = child.find(".//a:srgbClr", _DRAWINGML_NS)
        sysclr = child.find(".//a:sysClr", _DRAWINGML_NS)
        if srgb is not None and srgb.get("val"):
            colors.append(srgb.get("val", "000000").upper())
        elif sysclr is not None and sysclr.get("lastClr"):
            colors.append(sysclr.get("lastClr", "000000").upper())
        else:
            colors.append("000000")
    if len(colors) < 12:
        return _DEFAULT_THEME_RGB
    return tuple(colors[:12])


def _apply_tint(rgb_hex: str, tint: float) -> str:
    """Excel OOXML tint → RGB (MS Open XML spec)."""
    r = int(rgb_hex[0:2], 16) / 255.0
    g = int(rgb_hex[2:4], 16) / 255.0
    b = int(rgb_hex[4:6], 16) / 255.0

    def one(c: float) -> int:
        if tint < 0:
            c = c * (1.0 + tint)
        else:
            c = c * (1.0 - tint) + tint
        return max(0, min(255, int(round(c * 255))))

    return f"{one(r):02X}{one(g):02X}{one(b):02X}"


def _color_to_hex(color, theme_rgbs: tuple[str, ...]) -> str | None:
    if color is None:
        return None
    ctype = getattr(color, "type", None)
    if ctype == "rgb":
        rgb = getattr(color, "rgb", None)
        if not isinstance(rgb, str) or len(rgb) < 6:
            return None
        # openpyxl 자동색/투명: 00000000 만 무시 (00RRGGBB 는 유효한 색)
        if rgb.upper() in ("00000000", "None"):
            return None
        return rgb[-6:].upper()
    if ctype == "theme":
        idx = getattr(color, "theme", None)
        if idx is None or idx < 0 or idx >= len(theme_rgbs):
            return None
        base = theme_rgbs[idx]
        tint = float(getattr(color, "tint", 0.0) or 0.0)
        return _apply_tint(base, tint) if tint else base
    return None


def _cell_fill_hex(cell, theme_rgbs: tuple[str, ...]) -> str | None:
    fill = cell.fill
    if fill is None or getattr(fill, "patternType", None) != "solid":
        return None
    hex6 = _color_to_hex(fill.fgColor, theme_rgbs)
    if hex6 is None:
        return None
    if hex6 == "000000":
        # 순수 검정(theme0 tint0)은 배경으로 쓰지 않음. tint 있는 회색만 유지
        color = fill.fgColor
        if getattr(color, "type", None) != "theme":
            return None
        if float(getattr(color, "tint", 0.0) or 0.0) == 0.0:
            return None
    return f"#{hex6}"


def _cell_font_style(cell, theme_rgbs: tuple[str, ...]) -> str:
    font = cell.font
    parts: list[str] = []
    if font and font.bold:
        parts.append("font-weight:700")
    if font and font.color:
        hex6 = _color_to_hex(font.color, theme_rgbs)
        if hex6 and hex6 != "000000":
            parts.append(f"color:#{hex6}")
    return ";".join(parts)


def _format_cell_value(val: object) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        if val.hour or val.minute or val.second:
            return val.strftime("%Y-%m-%d %H:%M")
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, time):
        return val.strftime("%H:%M")
    text = str(val).replace("\x00", "")
    if len(text) > 200:
        text = text[:197] + "..."
    return text


def _merge_span_map(
    ws: Worksheet, max_row: int, max_col: int
) -> tuple[dict[tuple[int, int], tuple[int, int]], set[tuple[int, int]]]:
    """Return (top-left spans, covered non-origin cells) within crop."""
    spans: dict[tuple[int, int], tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()
    for rng in ws.merged_cells.ranges:
        min_r, min_c = rng.min_row, rng.min_col
        max_r, max_c = rng.max_row, rng.max_col
        if min_r > max_row or min_c > max_col:
            continue
        max_r = min(max_r, max_row)
        max_c = min(max_c, max_col)
        if min_r < 1 or min_c < 1:
            continue
        spans[(min_r, min_c)] = (max_r - min_r + 1, max_c - min_c + 1)
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) != (min_r, min_c):
                    covered.add((r, c))
    return spans, covered


def sheet_to_html(
    workbook_path: Path,
    sheet_name: str,
    *,
    max_row: int | None = None,
    max_col: int | None = None,
) -> str:
    """일자 시트를 엑셀에 가깝게 HTML로 렌더링 (하단 근무형태 블록 제외)."""
    with load_workbook_resilient(workbook_path, data_only=False) as wb:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"시트 없음: {sheet_name}")
        ws = wb[sheet_name]
        theme_rgbs = _parse_theme_rgbs(wb)
        end_row = find_report_crop_end_row(ws) if max_row is None else max_row
        end_col = max_col or min(ws.max_column or _PREVIEW_MAX_COL, _PREVIEW_MAX_COL)
        span_info, covered = _merge_span_map(ws, end_row, end_col)

        with load_workbook_resilient(workbook_path, data_only=True) as wb_vals:
            ws_vals = wb_vals[sheet_name]
            # 빈 행도 유지해야 병합/행높이 레이아웃이 엑셀과 비슷해짐
            rows_html: list[str] = []
            for r in range(1, end_row + 1):
                cells: list[str] = []
                row_h = ws.row_dimensions[r].height
                height_css = ""
                if row_h:
                    # Excel pt ≈ px for screen preview
                    height_css = f' style="height:{int(row_h)}px"'
                for c in range(1, end_col + 1):
                    if (r, c) in covered:
                        continue
                    cell = ws.cell(r, c)
                    val = ws_vals.cell(r, c).value
                    if val is None:
                        val = cell.value
                    text = _format_cell_value(val)
                    style_parts = [
                        "border:1px solid #b0b0b0",
                        "padding:3px 5px",
                        "font-size:11px",
                        "vertical-align:middle",
                        "white-space:pre-wrap",
                        "font-family:'Malgun Gothic',Calibri,sans-serif",
                    ]
                    bg = _cell_fill_hex(cell, theme_rgbs)
                    if bg:
                        style_parts.append(f"background:{bg}")
                    font_css = _cell_font_style(cell, theme_rgbs)
                    if font_css:
                        style_parts.append(font_css)
                    span_attrs = ""
                    if (r, c) in span_info:
                        rs, cs = span_info[(r, c)]
                        if rs > 1:
                            span_attrs += f' rowspan="{rs}"'
                        if cs > 1:
                            span_attrs += f' colspan="{cs}"'
                    cells.append(
                        f'<td{span_attrs} style="{";".join(style_parts)}">'
                        f"{escape(text)}</td>"
                    )
                rows_html.append(f"<tr{height_css}>" + "".join(cells) + "</tr>")

            body = "\n".join(rows_html) or "<tr><td>(빈 시트)</td></tr>"
            return (
                '<div style="background:#fff;color:#111;border:1px solid #999;'
                'border-radius:4px;overflow:auto;max-height:560px">'
                f'<div style="padding:8px 10px;font-weight:600;background:#e8f0fe;'
                f'border-bottom:1px solid #ccc">{escape(sheet_name)}'
                f' <span style="font-weight:400;color:#666;font-size:12px">'
                f"(첨부 범위 A1:{get_column_letter(end_col)}{end_row})</span></div>"
                '<table style="border-collapse:collapse;width:100%;table-layout:fixed">'
                f"{body}</table></div>"
            )


def report_crop_a1_range(
    ws: Worksheet, *, max_col: int | None = None
) -> tuple[str, int, int]:
    """첨부·미리보기용 A1 범위와 (end_row, end_col)을 반환."""
    end_row = find_report_crop_end_row(ws)
    end_col = max_col or min(ws.max_column or _PREVIEW_MAX_COL, _PREVIEW_MAX_COL)
    return f"A1:{get_column_letter(end_col)}{end_row}", end_row, end_col


def _trim_white_borders(png_path: Path) -> None:
    """PNG 가장자리 여백을 잘라 표만 남긴다."""
    try:
        from PIL import Image, ImageChops
    except ImportError:  # pragma: no cover
        return
    im = Image.open(png_path).convert("RGB")
    bg = Image.new("RGB", im.size, (255, 255, 255))
    diff = ImageChops.difference(im, bg)
    # 연한 회색 격자도 콘텐츠로 인식되도록 약간 증폭
    diff = ImageChops.add(diff, diff)
    bbox = diff.getbbox()
    if not bbox:
        return
    # 아주 작은 여백 유지
    pad = 4
    left = max(0, bbox[0] - pad)
    top = max(0, bbox[1] - pad)
    right = min(im.width, bbox[2] + pad)
    bottom = min(im.height, bbox[3] + pad)
    im.crop((left, top, right, bottom)).save(png_path)


def render_sheet_preview_png(
    workbook_path: Path,
    sheet_name: str,
    dest: Path,
    *,
    max_col: int | None = None,
) -> Path:
    """첨부용 crop 시트를 Excel→PDF→PNG로 렌더 (미리보기·메일용).

    로컬 Excel + pywin32 + pymupdf 필요. 실패 시 호출측에서 HTML 폴백.
    """
    try:
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
        import pymupdf
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "이미지 미리보기에는 pywin32와 pymupdf가 필요합니다."
        ) from exc

    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)

    # 첨부 파일과 동일한 crop xlsx를 만든 뒤 그 시트 전체를 PDF로 보냄
    cropped = dest.with_suffix(".crop.xlsx")
    pdf_path = dest.with_suffix(".pdf")
    export_sheet_workbook(workbook_path, sheet_name, cropped)

    # max_col은 export 단계에서 행 crop만 적용; PDF는 시트 전체(이미 crop됨)
    _ = max_col  # API 호환용

    # Streamlit 등 비메인 스레드에서는 COM 초기화가 필요
    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(
            str(cropped.resolve()),
            ReadOnly=True,
            UpdateLinks=0,
            IgnoreReadOnlyRecommended=True,
        )
        ws = wb.Worksheets(sheet_name)
        ws.Activate()
        # 인쇄 영역 = 사용 영역 (하단 근무형태는 이미 삭제됨)
        used = ws.UsedRange
        ws.PageSetup.PrintArea = used.Address
        ws.PageSetup.Zoom = False
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.FitToPagesTall = 1
        ws.PageSetup.LeftMargin = 0
        ws.PageSetup.RightMargin = 0
        ws.PageSetup.TopMargin = 0
        ws.PageSetup.BottomMargin = 0
        wb.ExportAsFixedFormat(0, str(pdf_path.resolve()))  # xlTypePDF
    finally:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:  # noqa: BLE001
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:  # noqa: BLE001
                pass
        pythoncom.CoUninitialize()

    if not pdf_path.is_file() or pdf_path.stat().st_size < 100:
        raise RuntimeError("Excel PDF 내보내기에 실패했습니다.")

    doc = pymupdf.open(pdf_path)
    try:
        page = doc[0]
        pix = page.get_pixmap(matrix=pymupdf.Matrix(2, 2), alpha=False)
        pix.save(str(dest))
    finally:
        doc.close()

    _trim_white_borders(dest)
    pdf_path.unlink(missing_ok=True)
    cropped.unlink(missing_ok=True)

    if not dest.is_file() or dest.stat().st_size < 100:
        raise RuntimeError("미리보기 PNG 생성에 실패했습니다.")
    return dest


def export_sheet_workbook(workbook_path: Path, sheet_name: str, dest: Path) -> Path:
    """해당 시트만 포함하고, 하단 근무형태 블록을 잘라낸 xlsx를 dest에 저장."""
    dest.parent.mkdir(parents=True, exist_ok=True)
    tmp = dest.with_suffix(".tmp.xlsx")
    # Excel이 원본을 잠근 경우에도 copy2는 대개 성공 → 복사본에서 가공
    try:
        shutil.copy2(workbook_path, tmp)
    except PermissionError as exc:
        raise PermissionError(
            f"엑셀 복사 실패(파일 잠금). Excel에서 저장·닫은 뒤 다시 시도하세요: {workbook_path}"
        ) from exc
    slim = openpyxl.load_workbook(tmp)
    try:
        if sheet_name not in slim.sheetnames:
            raise ValueError(f"시트 없음: {sheet_name}")
        for name in list(slim.sheetnames):
            if name != sheet_name:
                del slim[name]
        ws = slim[sheet_name]
        end_row = find_report_crop_end_row(ws)
        remerge: list[str] = []
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row > end_row:
                ws.unmerge_cells(str(rng))
            elif rng.max_row > end_row:
                ws.unmerge_cells(str(rng))
                top_left = f"{get_column_letter(rng.min_col)}{rng.min_row}"
                bottom_right = f"{get_column_letter(rng.max_col)}{end_row}"
                if top_left != bottom_right:
                    remerge.append(f"{top_left}:{bottom_right}")
        for ref in remerge:
            ws.merge_cells(ref)
        if ws.max_row and ws.max_row > end_row:
            ws.delete_rows(end_row + 1, ws.max_row - end_row)
        slim.save(tmp)
    finally:
        slim.close()
    only = openpyxl.load_workbook(tmp)
    try:
        only.save(dest)
    finally:
        only.close()
    tmp.unlink(missing_ok=True)
    return dest


def split_workday_slots(
    start: datetime, end: datetime, n: int
) -> list[tuple[datetime, datetime]]:
    """근무 구간을 n등분. 마지막 슬롯 end는 전체 end와 동일."""
    if n < 1:
        raise ValueError("n must be >= 1")
    if end <= start:
        raise ValueError("end must be after start")
    if n == 1:
        return [(start, end)]
    total = (end - start).total_seconds()
    step = total / n
    slots: list[tuple[datetime, datetime]] = []
    for i in range(n):
        s = start + timedelta(seconds=step * i)
        e = end if i == n - 1 else start + timedelta(seconds=step * (i + 1))
        slots.append((s, e))
    return slots


def find_time_overlaps(issues: list[FieldIssue]) -> list[str]:
    """포함된 이슈들의 Start/End 겹침 경고 메시지."""
    active = [
        i
        for i in issues
        if i.included and i.start is not None and i.end is not None and i.end > i.start
    ]
    warnings: list[str] = []
    for i, a in enumerate(active):
        for b in active[i + 1 :]:
            assert a.start is not None and a.end is not None
            assert b.start is not None and b.end is not None
            if a.start < b.end and b.start < a.end:
                warnings.append(
                    f"시간 겹침: Case {a.case_number} "
                    f"({a.start:%H:%M}–{a.end:%H:%M}) ↔ Case {b.case_number} "
                    f"({b.start:%H:%M}–{b.end:%H:%M})"
                )
    return warnings


def parse_field_issues(
    workbook_path: Path,
    sheet_name: str,
    *,
    day: date,
    fse_name: str,
    work_start: datetime,
    work_end: datetime,
) -> list[FieldIssue]:
    """Daily Note 영역에서 □ 이슈 (Case : ########) 줄을 파싱하고 시간을 균등 분할."""
    with load_workbook_resilient(workbook_path, data_only=True) as wb:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"시트 없음: {sheet_name}")
        ws = wb[sheet_name]
        end_row = find_report_crop_end_row(ws)
        lines: list[str] = []
        for r in range(23, end_row + 1):
            for c in range(2, min(ws.max_column or _PREVIEW_MAX_COL, _PREVIEW_MAX_COL) + 1):
                val = ws.cell(r, c).value
                if val is None:
                    continue
                text = str(val).replace("\x00", "").strip()
                if not text:
                    continue
                for part in re.split(r"[\r\n]+", text):
                    part = part.strip()
                    if part:
                        lines.append(part)
        # 이슈 줄 + 다음 이슈/섹션 헤더 전까지를 상세 본문으로 묶음
        stop_headers = {
            "next plan update",
            "data location",
            "data location:",
            "current problem statement",
            "current target statement",
            "daily field service note",
        }
        parsed: list[tuple[str, str, str]] = []
        i = 0
        while i < len(lines):
            line = lines[i]
            m = _ISSUE_LINE_RE.match(line)
            if not m:
                i += 1
                continue
            case_no = m.group("case")
            raw = line if line.lstrip().startswith(("□", "☐")) else f"□ {line.lstrip()}"
            detail_parts: list[str] = []
            j = i + 1
            while j < len(lines):
                nxt = lines[j]
                if _ISSUE_LINE_RE.match(nxt):
                    break
                if nxt.strip().lower() in stop_headers:
                    break
                if _CROP_END_MARKER in nxt:
                    break
                detail_parts.append(nxt)
                j += 1
            parsed.append((case_no, raw, "\n".join(detail_parts).strip()))
            i = j
        if not parsed:
            return []
        slots = split_workday_slots(work_start, work_end, len(parsed))
        issues: list[FieldIssue] = []
        for (case_no, raw, detail), (s, e) in zip(parsed, slots, strict=True):
            issues.append(
                FieldIssue(
                    case_number=case_no,
                    issue_line=raw,
                    activity_line=format_activity_line(day, fse_name, raw),
                    detail_text=detail,
                    start=s,
                    end=e,
                    included=True,
                )
            )
        return issues


def format_activity_line(day: date, fse_name: str, summary: str) -> str:
    """Case Activities__c 맨 위에 붙일 한 줄 요약."""
    name = fse_name.strip() or "미지정"
    summary = " ".join(summary.split())
    return f"{day:%Y-%m-%d} [{name}] {summary}"


def format_wo_description(
    issue: FieldIssue,
    *,
    asset_label: str,
    fse_name: str,
) -> str:
    """Work Order Description 용 상세 본문."""
    lines = [
        issue.issue_line,
        "",
        f"설비: {asset_label}",
        f"출장자: {fse_name.strip() or '미지정'}",
    ]
    if issue.start:
        lines.append(f"Start: {issue.start:%Y-%m-%d %H:%M}")
    if issue.end:
        lines.append(f"End: {issue.end:%Y-%m-%d %H:%M}")
    if issue.detail_text.strip():
        lines.extend(["", "[상세]", issue.detail_text.strip()])
    return "\n".join(lines)
